def conversion(textfile):
    # mypath should be the complete path for the directory containing the input text files

    def is_number(s):
        try:
            float(s)
            return True
        except ValueError:
            return False

    import xlwt

    style = xlwt.XFStyle()
    style.num_format_str = '#,###0.00'

    f = open(textfile, 'r+')
    row_list = []
    for row in f:
        row_list.append(row.split('|'))
    column_list = zip(*row_list)
    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet('Sheet1')
    i = 0
    for column in column_list:
        for item in range(len(column)):
            value = column[item].strip()
            if is_number(value):
                worksheet.write(item, i, float(value), style=style)
            else:
                worksheet.write(item, i, value)
        i += 1
    workbook.save(textfile.replace('.txt', '.xls'))


def array_work(file_path, level_setting, Environment_Name, Organization):
    import datetime
    import pandas
    import xlsxwriter
    import os
    from openpyxl import load_workbook
    df = pandas.read_excel(file_path, sheet_name='Sheet1')

    serv_def = str(df.values[0][0]).split('TION: ')[1]

    counting = 0
    project_count = 0
    start_index = 42
    while counting == 0:
        try:
            if str(df.values[start_index][0]) is None:
                counting = 1


            else:
                project_count = project_count + 1
                start_index = start_index + 48
        except IndexError:
            counting = 1

    serve_def_dict = {}
    for i in range(2, 40 + 1, 1):
        serve_def_dict[str(df.values[i][0]).split(',')[0]] = str(df.values[i][0]).split(',')[1]
        #print(i)

    projects_def_dict = {}
    project_start = 42
    for i in range(0, project_count, 1):
        projects_def_dict[str(df.values[project_start + (i * 48)][0]).split(': ')[1]] = {}
        for ii in range((44 + (i * 48)), (88 + 1 + (i * 48)), 1):
            projects_def_dict[str(df.values[project_start + (i * 48)][0]).split(': ')[1]][
                str(df.values[ii][0]).split(',')[0]] = \
                str(df.values[ii][0]).split(',')[1]

    outputfile = file_path.split('.')[0] + '.xlsx'
    if os.path.exists(outputfile):

       # print('must append instead')
        config_workboook = load_workbook(outputfile)

        active_sheet = config_workboook.active



        if str(active_sheet['A1'].value) == 'Organization':
         #   print('yes right first cell')

            if active_sheet.max_column == 9:
             #   print('yes right amount of columns')

                append_start=active_sheet.max_row+1;

               # print('start addding at row: '+str(append_start))


                row = append_start
                row_id=1
                col = 7+1
                index = 0
                # Iterate over the data and write it out row by row.



                current_time = datetime.datetime.now()
                # write the server def details
                for key, value in sorted(serve_def_dict.items()):
                    active_sheet.cell(row=row, column=1).value = Organization
                    active_sheet.cell(row=row, column=2).value = Environment_Name
                    active_sheet.cell(row=row, column=3).value = level_setting
                    active_sheet.cell(row=row, column=col).value = key
                    active_sheet.cell(row=row, column=col+1).value = value
                    active_sheet.cell(row=row, column=7).value = row_id
                    active_sheet.cell(row=row, column=4).value = serv_def
                    active_sheet.cell(row=row, column=5).value = 'Server Configuration'
                    active_sheet.cell(row=row, column=6).value = current_time
                    active_sheet.cell(row=row, column=6).number_format = 'mmm d yyyy hh:mm AM/PM'

                    row += 1
                    row_id+= 1
                # print(row)
                # write the project def details
                for i in range(0, project_count, 1):
                    project_desc_id = 40;
                    for key, value in sorted(projects_def_dict.items())[1][1].items():
                        active_sheet.cell(row=row, column=1).value = Organization
                        active_sheet.cell(row=row, column=2).value = Environment_Name
                        active_sheet.cell(row=row, column=3).value = level_setting
                        active_sheet.cell(row=row, column=col).value = key
                        active_sheet.cell(row=row, column=col + 1).value = value
                        active_sheet.cell(row=row, column=7).value = project_desc_id
                        active_sheet.cell(row=row, column=4).value = serv_def
                        active_sheet.cell(row=row, column=5).value = str(sorted(projects_def_dict.items())[i][0])
                        active_sheet.cell(row=row, column=6).value = current_time
                        active_sheet.cell(row=row, column=6).number_format = 'mmm d yyyy hh:mm AM/PM'

                        project_desc_id = project_desc_id + 1;
                        row += 1
                config_workboook.save(outputfile)
    else:
        workbook = xlsxwriter.Workbook(outputfile)
        worksheet = workbook.add_worksheet()

        # Start from the first cell. Rows and columns are zero indexed.
        row = 1
        col = 7
        worksheet.write(0, 0, 'Organization')
        worksheet.write(0, 1, 'Environment Name')
        worksheet.write(0, 2, 'Level')
        worksheet.write(0, 3, 'Server Def')
        worksheet.write(0, 4, 'Project Name')
        worksheet.write(0, 6, 'Desc ID')
        worksheet.write(0, 7, 'Desc')
        worksheet.write(0, 8, 'Value')
        worksheet.write(0, 5, 'Time Collected')
        index = 0
        # Iterate over the data and write it out row by row.

        format7 = workbook.add_format({'num_format': 'mmm d yyyy hh:mm AM/PM'})

        current_time = datetime.datetime.now()
        # write the server def details
        for key, value in sorted(serve_def_dict.items()):
            worksheet.write(row, 0, Organization)
            worksheet.write(row, 1, Environment_Name)
            worksheet.write(row, 2, level_setting)
            worksheet.write(row, col, key)
            worksheet.write(row, col + 1, value)
            worksheet.write(row, 6, row)
            worksheet.write(row, 3, serv_def)
            worksheet.write(row, 4, 'Server Configuration')
            worksheet.write(row, 5, current_time, format7)
            row += 1

        # print(row)
        # write the project def details
        for i in range(0, project_count, 1):
            project_desc_id = 40;
            for key, value in sorted(projects_def_dict.items())[1][1].items():
                worksheet.write(row, 0, Organization)
                worksheet.write(row, 1, Environment_Name)
                worksheet.write(row, 2, level_setting)
                worksheet.write(row, col, key)
                worksheet.write(row, col + 1, value)
                worksheet.write(row, 3, serv_def)
                worksheet.write(row, 4, str(sorted(projects_def_dict.items())[i][0]))
                worksheet.write(row, 6, project_desc_id)
                worksheet.write(row, 5, current_time, format7)
                project_desc_id = project_desc_id + 1;
                row += 1

        workbook.close()
    os.remove(file_path)

